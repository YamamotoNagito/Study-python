import datetime
import os
import openpyxl

today = datetime.datetime.today()

year = today.year
month = today.month
day = today.day
hour = today.hour
hour = today.hour
minute = today.minute
second = today.second
microsecond = today.microsecond

print(year)  # 2021
print(month)  # 4
print(day)  # 24
print(hour) # 11
print(minute) # 21

# ファイルの有無を確認してくれるためのコマンド osのimportが必要になってる
# print("hello.txt : ",os.path.exists('hello.txt'))
# print('./excel/Sample.xlsx',os.path.exists('./excel/Sample.xlsx'))

path = "./excel/" + str(year) + "-" + str(month) + ".xlsx"
if os.path.exists(path) == False:
    print("当月のファイルがありませんので作成します")
    
    wb = openpyxl.Workbook() # エクセルファイルの新規作成
    print("excelファイルの作成が完了しました")
    wb.remove(wb.worksheets[0]) # 最初のシートの削除

else:
    wb = openpyxl.load_workbook(path)
    print("既に作成されています")

sheetname = str(year) + "." + str(month) + "-" + str(day) + "." + str(hour) + "-" + str(minute)
ws = wb.create_sheet(title=sheetname)
print(wb.sheetnames)

# 9×9表の作成
# for i in range(10):
#     for j in range(10):
#         ws.cell(i+1,j+1).value = i*j

# 画像の保存
PhotoPath = "./image/623.png"
img = openpyxl.drawing.image.Image(PhotoPath)
# 本来はimg,"A3"という風に書くがその代わりにほかの方法で書く(sheet.cell(row=行番号,column=列番号).coordinate)
ws.add_image(img, ws.cell(row=3,column=3).coordinate) #行・列の指定の際には変数を用いてもよい

PhotoPath = "./image/623-1.png"
img = openpyxl.drawing.image.Image(PhotoPath)
ws.add_image(img, "D3")


# セルのサイズの変更
ws.row_dimensions[3].height = 175       # 列の高さを変更
ws.column_dimensions['C'].width = 50   # 列の幅を変更
ws.column_dimensions['D'].width = 50   # 列の幅を変更
ws.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 50   # 列の幅を変更 Eの代わりに5番目の列ということを指定する

# セルの幅を調整する
# for col in ws.columns:
#     max_length = 0
#     column = col[0].column

#     for cell in col:
#         if len(str(cell.value)) > max_length:
#             max_length = len(str(cell.value))

#     adjusted_width = (max_length + 2) * 1.2
#     ws.column_dimensions[column].width = adjusted_width

# 上書き保存（読み込んだのと同じ名前を指定）
wb.save(path) # エクセルファイルの保存