import openpyxl

# エクセルファイルの読み込み
wb = openpyxl.load_workbook("./excel/Sample.xlsx")

# エクセルファイルの新規作成
# wb = openpyxl.Workbook()

# 別名保存（読み込んだのと別名を指定）
# wb.save("Sample2.xlsx")

# 名前が「Sheet」のシート
ws = wb["Sheet4"]
print("変更前 : " + ws.title)
print("シート一覧 : ")
print(wb.sheetnames)

# 先頭のシート（注意：インデックス番号は0から始まる）
# ws = wb.worksheets[0]

# インデックス番号の確認
# wb.index(ws)

# シート名の変更
ws.title = "SheetChange"
print("シート一覧 : ")
print(wb.sheetnames)

# 「Sheet4」を末尾に追加
ws4 = wb.create_sheet(title="Sheet4")
print(wb.sheetnames)

# 「New Sheet」を左から３つ目の位置に追加
# ws_new = wb.create_sheet(title="New Sheet", index=2)
# print(wb.sheetnames)

# 「Sheet2」のコピーをシートの最後に追加
# ws2_copy = wb.copy_worksheet(wb["Sheet4"])
# print(wb.sheetnames)
# コピーしたシートは末尾に「 Copy」が付く

# 「Sheet2」のコピーを削除
# wb.remove(ws2_copy)
# print(wb.sheetnames)

# 最初のシートを削除
wb.remove(wb.worksheets[0])
# wb.remove(wb.worksheets[-1]) 末尾のシートを削除
print(wb.sheetnames)


# エクセルファイルの保存
# 上書き保存（読み込んだのと同じ名前を指定）
wb.save("./excel/Sample.xlsx")