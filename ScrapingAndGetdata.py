import datetime
import os
import openpyxl
import sys
from selenium import webdriver
from time import sleep
from PIL import Image
import shutil

today = datetime.datetime.today()

year = today.year
month = today.month
day = today.day
hour = today.hour
hour = today.hour
minute = today.minute
second = today.second
microsecond = today.microsecond

filepath = "./excel/" + str(year) + "-" + str(month) + ".xlsx"
filepathcopy = "./excel/" + str(year) + "-" + str(month) + "-copy.xlsx"
if os.path.exists(filepath) == False:
    print("当月のファイルがありませんので作成します")
    
    wb = openpyxl.Workbook() # エクセルファイルの新規作成
    print("excelファイルの作成が完了しました")
    wb.remove(wb.worksheets[0]) # 最初のシートの削除

else:
    wb = openpyxl.load_workbook(filepath)
    print("既に作成されています")

sheetname = str(year) + "." + str(month) + "-" + str(day) + "." + str(hour) + "-" + str(minute)
ws = wb.create_sheet(index=0,title=sheetname)
print(wb.sheetnames)

driver = webdriver.Chrome(executable_path='./chromedriver') #絶対パス指定も可能

# # ページへの遷移
url = "url名" # url名を必要に応じて入力
driver.get(url)

# class属性を探すようにしていく urlの取得
class_name = "class名" #class名を必要に応じて入力 
class_elems = driver.find_elements_by_class_name(class_name)
class_elems_a = []
for elem in class_elems: #class属性を持つものすべてについて適応
    class_elems_a.append(elem.find_element_by_tag_name("a").get_attribute('href'))

# エクセルの列・行を指定
row = 3
col = 3 

# # 一覧を取得していくための部分
for elem_a in class_elems_a:
    number_url = elem_a
    driver.get(number_url)
    class_name2 = "btn-base"
    class_elems2 = driver.find_elements_by_class_name(class_name2)
    for elem2 in class_elems2: #class属性を持つものすべてについて適応

        if elem2.text.isdigit(): #番号の有無を判定 ボタンとして「検索」などの部分をはじいていく必要がある
            print(elem2.text) #番号の取得

            # セルに値の代入を行っていく
            ws.cell(row,col).value = elem2.text

            # File Name
            p1 = "./image/"
            p2 = elem2.text
            p3 = ".png"
            path = f"{p1}{p2}{p3}"
            print(path)

            url = "url名"+elem2.text #url名を必要に応じて入力
            driver.execute_script("window.open(arguments[0], '_blank')", url)
            driver.switch_to.window(driver.window_handles[1]) # 新しいタブへの切り替え
            
            FILENAME = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
            driver.set_window_size(750,1000) # window size をセットする
            driver.save_screenshot(FILENAME) # Screen Shot を行う
            
            # 画像のリサイズの一連の作業
            img_view = Image.open(path) #画像の読み込み
            img_view = img_view.crop((0, 100, 750, 600)) # 画像のトリミング(left,top,rigt,under)→順番注意
            small_img_view = img_view.resize((round(img_view.width * 0.5), round(img_view.height * 0.4))) #画像のリサイズ
            small_img_view.save(path) #リサイズした画像を名前をつけて保存
           
            # 本来はimg,"A3"という風に書くがその代わりにほかの方法で書く(sheet.cell(row=行番号,column=列番号).coordinate)
            PhotoPath = path
            img = openpyxl.drawing.image.Image(PhotoPath)
            ws.add_image(img, ws.cell(row=row,column=col+1).coordinate) #行・列の指定の際には変数を用いてもよい
            
            # DeletePath1 = PhotoPath #削除用のファイルのパス

            # セルのサイズの変更
            ws.row_dimensions[row].height = 175       # 列の高さを変更
            ws.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 50   # 列の幅を変更 Dの代わりに5番目の列ということを指定する 改良余地あり
            ws.column_dimensions[openpyxl.utils.get_column_letter(col + 2)].width = 50   # 列の幅を変更 Eの代わりに5番目の列ということを指定する

            # 2枚目のスクショ(データの写真)を保存するように p2はそのまま使用する
            p1 = "./image/"
            p3 = "-1.png"
            path = f"{p1}{p2}{p3}"
            driver.execute_script("window.scrollTo(0, 615);") # スクロールする
            sleep(5.0) 
            FILENAME = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
            driver.set_window_size(750,1000) # window size のセット
            driver.save_screenshot(FILENAME) # Screen Shot を行う

            
            # 画像のリサイズの一連の作業
            x1 = 120 # 取得する画像のサイズを指定
            y1 = 120 # 取得する画像のサイズを指定
            x2 = 120 # 取得する画像のサイズを指定
            y2 = 120 # 取得する画像のサイズを指定            
            img_view = Image.open(path) #画像の読み込み
            img_view = img_view.crop((x1, y1, x2, y2)) # 画像のトリミング
            small_img_view = img_view.resize((round(img_view.width * 0.5), round(img_view.height * 0.4))) #画像のリサイズ
            small_img_view.save(path) #リサイズした画像を名前をつけて保存
            # 本来はimg,"A3"という風に書くがその代わりにほかの方法で書く(sheet.cell(row=行番号,column=列番号).coordinate)
            PhotoPath = path
            img = openpyxl.drawing.image.Image(PhotoPath)
            ws.add_image(img, ws.cell(row=row,column=col+2).coordinate) #行・列の指定の際には変数を用いてもよい

            # DeletePath2 = PhotoPath #削除用のファイルのパス

            # 開いているタブを閉じる
            # sleep(1)

            driver.close()
            driver.switch_to.window(driver.window_handles[0]) #switch tab

            row = row + 1
            # 画像の削除
            # os.remove(DeletePath1)
            # os.remove(DeletePath2)

sleep(3)
driver.close()
driver.quit()

wb.save(filepath) # エクセルファイルの保存
shutil.copyfile(filepath,filepathcopy) # ファイルのコピー(excel破損や操作時のバックアップ用)